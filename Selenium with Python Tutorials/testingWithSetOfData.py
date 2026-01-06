import time
from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By

#Load the Excel Sheet
workbook = load_workbook('testdata.xlsm')
sheet = workbook.active

chrome_options = Options()
chrome_options.add_argument("--incognito")
browser = webdriver.Chrome(options=chrome_options)
browser.maximize_window()
url = "https://www.saucedemo.com/"
browser.get(url)

for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
    username = row[0]
    password = row[1]

    browser.find_element(By.ID, "user-name").send_keys(username)
    browser.find_element(By.ID, "password").send_keys(password)
    browser.find_element(By.ID, "login-button").click()
    time.sleep(5)
    browser.find_element(By.ID, "react-burger-menu-btn").click()
    time.sleep(5)
    browser.find_element(By.ID, "logout_sidebar_link").click()

browser.quit()

